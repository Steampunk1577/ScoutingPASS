from fileinput import filename
import cv2
from cv2 import imshow
import numpy as np
from pyzbar.pyzbar import decode
import openpyxl
import QRscanner

fileName = "Excel/Book1.xlsx"
offset = 1


def column_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


def formatData(rawData):
    data = rawData.split(';')
    formatedData = []
    for i in range(len(data)):
        splited = data[i].split('=')
        formatedData.append(splited[1])
    
    #formatData is a list full of strings, we will format int and lists to be int and lists.
    for i in range(len(formatedData)):
        value = formatedData[i]
        if value.isnumeric():
                value = int(value)
        formatedData[i] = value
    return formatedData


def ExcelHandler(data):
    formatedData = formatData(data)
    excelWorkbook = openpyxl.load_workbook(filename=fileName)
    excelSheet = excelWorkbook["raw"]
    maxRow = excelSheet.max_row
    #print(maxRow)

    for line in range(1, maxRow + 1):
        duplicate = True
        for col, val in enumerate(formatedData, start=1):
            duplicate &= (excelSheet.cell(row=line, column=col).value == val) \
                          or (excelSheet.cell(row=line, column=col).value is None and val == "")
            if not duplicate:
                break
        if duplicate:
            print("This line already exists")
            return

    for col, d in enumerate(formatedData, start=1):
        excelSheet.cell(row=maxRow + 1, column=col).value = d
    excelWorkbook.save(fileName)


if __name__ == "__main__":
    cap = cv2.VideoCapture(0)
    data = None
    while True:
        while True:
            valid, frame = cap.read()
            data = QRscanner.read_barcodes(frame)
            cv2.imshow("QR Scanner", frame)
            key = cv2.waitKey(1)
            if key == 32:  # Spacebar
                cv2.destroyAllWindows()
                break  # esc to quit
            elif key == 27:  # ESC
                cv2.destroyAllWindows()
                print("Exiting because ESC pressed")
                exit(0)

        #for testing purposes
        #data = "s=Ofir;e=2022isde1;l=qm;m=5;r=b2;t=8223;as=[31];at=Y;au=2;al=1;am=3;tu=3;tmh=1;tl=2;wd=Y;ss=[18,32,55,53,28];c=3;lsr=2;be=Y;ds=3;dr=5;ba=Y;d=N;cf=Y;all=N;co=yess;cnf=a"
        if data[1] != "":
            ExcelHandler(data[1])

    



